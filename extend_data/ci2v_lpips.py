# Compare an image with every frame of a video to find the best match

import os
import operator
import time
import datetime
import warnings
import cv2
import torch
import openpyxl 




from sys import stdout
#from itertools import izip

from skimage import color
from skimage.metrics import structural_similarity as ssim
import lpips

import os
os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE"

#record start time
start = time.time()

#ignore non-contiguous skimage warning
warnings.filterwarnings("ignore", module="skimage")


def prepare_image(filename):
    #open still image as rgb
    img = cv2.imread(filename, cv2.IMREAD_COLOR)
    #shrink
    img = cv2.resize(img, (256, 256))
    #convert to b&w
    #img = color.rgb2gray(img)
    return img


def best_match(similarities):
    d = max(similarities, key=lambda x:x['similarity'])
    best_frame_number = d['frame']
    best_similarity = d['similarity']
    return best_frame_number, best_similarity


def parse_video(image, video, n_matches, break_point=False, verbose=False, metric=None):
    #iterate through video frames
    
    similarities = [{'frame': 0, 'similarity': 0}]
    frame_count = 0
    
    #get current time
    fps_time = time.time()

    cap = cv2.VideoCapture(video)
    while(cap.isOpened()):

        ret, frame = cap.read()

        #break at EOF
        if (type(frame) == type(None)):
            cap.release()
            break

        #increment frame counter
        frame_count += 1

        if frame_count%5 == 1:
            #resize current video frame
            small_frame = cv2.resize(frame, (256, 256))
            #convert to greyscale
            #small_frame_bw = color.rgb2gray(small_frame)

            #compare current frame to source image
            #lpips_loss = lpips.LPIPS(net='alex')
            similarity = metric(torch.from_numpy(image).permute(2,0,1), torch.from_numpy(small_frame).permute(2,0,1)).item()
            #similarity = ssim(image, small_frame_bw, win_size = 5, data_range= small_frame_bw.max() -  small_frame_bw.min())

            #remember current frame details
            similarities.append({'frame'      : frame_count,
                                'similarity' : similarity,
                                'image'      : frame})

            #find best match overall
            best_frame_number, best_similarity = best_match(similarities)
            
            #sort similarities list
            similarities = sorted(similarities, key=operator.itemgetter('similarity'), reverse=False)
            #remove surplus entries
            similarities = similarities[:n_matches]

            #calculate fps
            fps = frame_count / (time.time() - fps_time)

            #feedback to cli
            stdout.write('\r@ %d [%sfps] | best: %d (%s)  \r'
                % (frame_count, int(round(fps)), best_frame_number, round(best_similarity, 4), ))
            stdout.flush()

            #handle break_point
            if break_point:
                if similarity >= break_point:
                    return similarities

    cap.release()
    return similarities


def sort_results(results, output=False):
    #sort results
    print('\n')
    sorted_results = sorted(results, key=operator.itemgetter('similarity'), reverse=False)
    n = 0
    print('\n--results:')
    for res in sorted_results:
        n += 1
        print ('#%s\t%s\t%s\t: %s' % (n, res['filename'], res['frame'], res['similarity']))

        #save matched frames
        if output:
            save_frame(output, n, res['image'])


def save_frame(filename, n, image):
    fn, ext = filename.split('.')
    fn = '%s_%s.%s' % (fn, n, ext)
    cv2.imwrite(fn, image)


def walk(source_image, directory, number=1, break_point=False):
    results = []
    extentions = ['mp4', 'avi', 'mov', 'mkv', 'm4v']
    for root, dirs, files in os.walk(directory):
        for file in files:
            for ext in extentions:
                if file.endswith(ext):
                    video_fn = (os.path.join(root, file))
                    print (video_fn)
                    similarities = parse_video(source_image,
                                               video_fn,
                                               n_matches=number,
                                               break_point=break_point)
                     
                    #flatten results
                    for d in similarities:
                        results.append({'filename'   : video_fn,
                                        'frame'      : d['frame'],
                                        'similarity' : d['similarity'],
                                        'image'      : d['image']})

                        #stop walk if break point achieved
                        if break_point:
                            if d['similarity'] >= break_point:
                                return results

    return results

def write_exel(workbook, value1, value2, value3):

    # Select the active worksheet (or specify the sheet name)
    sheet = workbook.active

    # Find the first empty row
    empty_row = sheet.max_row + 1

    # Define the new data for the new line
    new_data = [value1, value2, value3]

    # Write data to the first empty row
    for col_num, value in enumerate(new_data, start=1):
        sheet.cell(row=empty_row, column=col_num).value = value

def main():
    import argparse

    #define cli arguments
    parser = argparse.ArgumentParser(description='''
        Compare an image with every frame of a video
        to find the best match.

        ============================================
        Edward Anderson
        --------------------------------------------
        v0.1 | 2016
        ''', formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('-i', '--image', help='source image')
    parser.add_argument('-v', '--video', help='video to search inside')
    parser.add_argument('-n', '--number', help='number of best matches to return', type=int, default=1)
    parser.add_argument('-b', '--break_point', help='stop searching when frame with [break_point] accuracy found; a number between 0 and 1', type=float, default=False)
    parser.add_argument('-o', '--output', help='filename.ext for best match; saved files are appended with "_n.ext"')
    parser.add_argument('-d', '--directory', help='directory of videos')
    args = parser.parse_args()

    #check source and destination provided
    if not args.image:
        parser.error('argument -i / --image is required')
    if not args.video:
        if not args.directory:
            parser.error('argument -v / --video is required')

    #prepare image
    source_image = prepare_image(args.image)
    lpips_loss = lpips.LPIPS(net='vgg')

    # Load the existing workbook
    excel_path = "extended_image_list_lpips.xlsx"
    workbook = openpyxl.load_workbook(excel_path)

    #either walk directory or hande single file
    if args.directory:
        #scan directory and process each video file
        print ('\n--reading videos:')
        results = walk(source_image, args.directory, args.number, args.break_point)
        s_results = sort_results(results, args.output)
        
    else:
        #process single video file
        print ('\n--reading video:')
        similarities = parse_video(source_image,
                                   args.video,
                                   n_matches=args.number,
                                   break_point=args.break_point,
                                   metric = lpips_loss)

        print ('\n\n--results:')
        #results to cli
        n = 0
        for d in similarities:
            if 'image' in d:
                n += 1
                print ('#%s\t%s\t: %s' % (n, d['frame'], d['similarity']))
                
                #save matched frames
                if args.output:
                        save_frame(args.output, n, d['image'])
                write_exel(workbook, args.output[:-4]+f'_{n}'+'.jpg',d['frame'], d['similarity'])

    
    # Save the workbook
    workbook.save(excel_path)

    seconds_taken = time.time() - start
    time_taken = str(datetime.timedelta(seconds=seconds_taken))
    print ('\n--time taken: \n%s\n' % time_taken)


if __name__ == '__main__':
    main()



    


